"""Print section header rows for all sheets so we know what keywords to add."""
import openpyxl, sys

EXCEL = r"C:\Users\LATIB PRO\Downloads\download (1).xlsx"
wb = openpyxl.load_workbook(EXCEL, read_only=True, data_only=True)

for sheet_name in wb.sheetnames:
    if sheet_name.lower() == "metadata":
        continue
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    print(f"\n=== Sheet: {sheet_name!r} ===")
    for i, row in enumerate(rows):
        first = row[0]
        if not first or not isinstance(first, str) or not first.strip():
            continue
        next_row = rows[i+1] if i+1 < len(rows) else None
        if next_row:
            nxt = str(next_row[0]).strip() if next_row[0] else ""
            print(f"  Row {i:4d}: title={str(first)[:60]!r:62s}  next={nxt!r}")
