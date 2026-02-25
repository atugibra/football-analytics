"""
Football Data Excel Importer
Reads download (1).xlsx and imports all data into the PostgreSQL database.
Usage: python import_excel.py --file "path/to/download (1).xlsx"
"""

import os, re, json, argparse
import openpyxl
import psycopg2
from psycopg2.extras import RealDictCursor
from dotenv import load_dotenv

load_dotenv(dotenv_path=os.path.join(os.path.dirname(__file__), "../api/.env"))

DATABASE_URL = os.getenv("DATABASE_URL")

# ─── DB helpers ──────────────────────────────────────────────────────────────
def connect():
    return psycopg2.connect(DATABASE_URL, cursor_factory=RealDictCursor)

def get_or_create_id(cur, table, unique_cols: dict, extra_cols: dict = {}):
    where = " AND ".join(f"{k}=%s" for k in unique_cols)
    cur.execute(f"SELECT id FROM {table} WHERE {where}", list(unique_cols.values()))
    row = cur.fetchone()
    if row:
        return row["id"]
    all_cols = {**unique_cols, **extra_cols}
    cols   = ", ".join(all_cols.keys())
    ph     = ", ".join(["%s"] * len(all_cols))
    cur.execute(f"INSERT INTO {table} ({cols}) VALUES ({ph}) RETURNING id", list(all_cols.values()))
    return cur.fetchone()["id"]

def safe(v):
    if v is None: return None
    s = str(v).strip()
    return None if s in ("", "nan", "None", "N/A") else s

def safe_int(v):
    try: return int(str(v).strip())
    except: return None

def safe_float(v):
    try: return float(str(v).strip())
    except: return None

def parse_date(v):
    s = safe(v)
    if not s: return None
    s = s.replace("-", "")
    if len(s) == 8 and s.isdigit():
        return f"{s[:4]}-{s[4:6]}-{s[6:]}"
    return s[:10] if len(s) >= 10 else None

def parse_score(v):
    s = safe(v)
    if not s: return None, None
    cleaned = re.sub(r"\s*\(.*?\)", "", s).strip()
    for sep in ["–", "-", "—"]:
        if sep in cleaned:
            parts = cleaned.split(sep)
            try: return int(parts[0].strip()), int(parts[1].strip())
            except: return None, None
    return None, None

# ─── Sheet section parser ──────────────────────────────────────────────────
SECTION_TYPES = {
    "SCORES & FIXTURES": "fixtures",
    "SCORES &amp; FIXTURES": "fixtures",
    "SQUAD STANDARD STATS": "standard",
    "SQUAD GOALKEEPING": "goalkeeping",
    "SQUAD SHOOTING": "shooting",
    "SQUAD PLAYING TIME": "playing_time",
    "SQUAD MISCELLANEOUS": "misc",
    "PLAYER STANDARD STATS": "player",
    "NATIONALITIES": "skip",
    "TABLE": "standings",
}

def detect_section_type(title: str):
    t = title.upper()
    for key, stype in SECTION_TYPES.items():
        if key in t:
            return stype
    return "skip"

def parse_season_from_title(title: str):
    m = re.search(r"(\d{4}-\d{4})", title)
    return m.group(1) if m else None

def parse_league_from_title(title: str):
    m = re.search(r"\d{4}-\d{4}\s+(.+?)\s+TABLE", title, re.IGNORECASE)
    return m.group(1).strip() if m else None

def rows_from_sheet(ws):
    return list(ws.iter_rows(values_only=True))

def find_sections(rows):
    sections = []
    for i, row in enumerate(rows):
        first = row[0]
        if not first or not isinstance(first, str) or not first.strip(): continue
        next_row = rows[i+1] if i+1 < len(rows) else None
        if next_row and next_row[0] and isinstance(next_row[0], str):
            nxt = next_row[0].strip().lower()
            if nxt in ("team", "ranker", "gameweek", "rank", "round", "dayofweek"):
                sections.append(i)
    return sections

def section_rows(rows, sections, idx):
    start = sections[idx]
    end   = sections[idx+1] if idx+1 < len(sections) else len(rows)
    title   = str(rows[start][0]).strip()
    headers = [str(h).strip() if h is not None else "" for h in rows[start+1]]
    data    = [row for row in rows[start+2:end] if any(c is not None for c in row)]
    return title, headers, data

# ─── Import functions ──────────────────────────────────────────────────────
def import_fixtures(cur, league_id, season_id, headers, data):
    h = headers
    inserted = 0
    for row in data:
        d = dict(zip(h, row))
        home = safe(d.get("home_team")); away = safe(d.get("away_team"))
        if not home or not away: continue
        home_id = get_or_create_id(cur, "teams", {"name": home, "league_id": league_id})
        away_id = get_or_create_id(cur, "teams", {"name": away, "league_id": league_id})
        hs, as_ = parse_score(d.get("score"))
        cur.execute("""
            INSERT INTO matches (league_id, season_id, home_team_id, away_team_id,
                gameweek, dayofweek, match_date, start_time, home_score, away_score,
                score_raw, attendance, venue, referee, round)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            ON CONFLICT (home_team_id, away_team_id, match_date) DO UPDATE SET
                home_score=EXCLUDED.home_score, away_score=EXCLUDED.away_score,
                score_raw=EXCLUDED.score_raw, attendance=EXCLUDED.attendance, updated_at=NOW()
        """, (
            league_id, season_id, home_id, away_id,
            safe_int(d.get("gameweek")), safe_int(d.get("dayofweek")),
            parse_date(d.get("date")), safe(d.get("start_time")),
            hs, as_, safe(d.get("score")),
            safe_int(d.get("attendance")), safe(d.get("venue")),
            safe(d.get("referee")), safe(d.get("round"))
        ))
        inserted += 1
    return inserted

def import_squad_stats(cur, league_id, season_id, stat_type, headers, data):
    inserted = 0
    for row in data:
        d = dict(zip(headers, row))
        team_raw = safe(d.get("team"))
        if not team_raw: continue
        split     = "against" if team_raw.startswith("vs ") else "for"
        team_name = team_raw[3:] if split == "against" else team_raw
        team_id   = get_or_create_id(cur, "teams", {"name": team_name, "league_id": league_id})
        payload   = {k: safe(v) for k, v in d.items() if k != "team"}
        cur.execute("""
            INSERT INTO team_squad_stats
                (team_id, league_id, season_id, split, players_used, avg_age, possession,
                 games, games_starts, minutes, minutes_90s, goals, assists,
                 standard_stats, goalkeeping, shooting, playing_time, misc_stats)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            ON CONFLICT (team_id, season_id, split) DO UPDATE SET
                scraped_at=NOW()
        """, (
            team_id, league_id, season_id, split,
            safe_int(d.get("players_used")), safe_float(d.get("avg_age")),
            safe_float(d.get("possession")), safe_int(d.get("games")),
            safe_int(d.get("games_starts")), safe_int(d.get("minutes")),
            safe_float(d.get("minutes_90s")), safe_int(d.get("goals")),
            safe_int(d.get("assists")),
            json.dumps(payload) if stat_type == "standard" else "{}",
            json.dumps(payload) if stat_type == "goalkeeping" else "{}",
            json.dumps(payload) if stat_type == "shooting" else "{}",
            json.dumps(payload) if stat_type == "playing_time" else "{}",
            json.dumps(payload) if stat_type == "misc" else "{}",
        ))
        inserted += 1
    return inserted

def import_players(cur, season_id, league_id, headers, data):
    inserted = 0
    for row in data:
        d = dict(zip(headers, row))
        name = safe(d.get("player"))
        if not name or name.lower() == "player": continue
        team_name = safe(d.get("team"))
        team_id   = get_or_create_id(cur, "teams", {"name": team_name, "league_id": league_id}) if team_name else None
        payload   = {k: safe(v) for k, v in d.items()}
        cur.execute("""
            INSERT INTO player_stats
                (player_name, nationality, position, team_id, season_id,
                 age, birth_year, games, games_starts, minutes, minutes_90s,
                 goals, assists, standard_stats)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            ON CONFLICT (player_name, team_id, season_id) DO UPDATE SET
                goals=EXCLUDED.goals, assists=EXCLUDED.assists,
                standard_stats=EXCLUDED.standard_stats, scraped_at=NOW()
        """, (
            name, safe(d.get("nationality")), safe(d.get("position")),
            team_id, season_id,
            safe_int(d.get("age")), safe_int(d.get("birth_year")),
            safe_int(d.get("games")), safe_int(d.get("games_starts")),
            safe_int(d.get("minutes")), safe_float(d.get("minutes_90s")),
            safe_int(d.get("goals")), safe_int(d.get("assists")),
            json.dumps(payload)
        ))
        inserted += 1
    return inserted

# ─── Main ─────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--file", required=True, help="Path to the Excel file")
    args = parser.parse_args()

    print(f"🔌 Connecting to database...")
    conn = connect()
    cur  = conn.cursor()

    print(f"📂 Opening {args.file}")
    wb = openpyxl.load_workbook(args.file, read_only=True, data_only=True)

    total_fixtures = total_squad = total_players = 0

    for sheet_name in wb.sheetnames:
        if sheet_name.lower() == "metadata":
            print(f"  ⏭  Skipping Metadata sheet")
            continue

        ws   = wb[sheet_name]
        rows = rows_from_sheet(ws)
        secs = find_sections(rows)

        print(f"\n📋 Sheet: {sheet_name!r} | {len(secs)} sections")

        for i in range(len(secs)):
            title, headers, data = section_rows(rows, secs, i)
            stype  = detect_section_type(title)
            season = parse_season_from_title(title)
            league = parse_league_from_title(title)

            if not league or not season or stype == "skip":
                continue

            league_id = get_or_create_id(cur, "leagues", {"name": league})
            season_id = get_or_create_id(cur, "seasons", {"name": season})

            if stype == "fixtures":
                n = import_fixtures(cur, league_id, season_id, headers, data)
                total_fixtures += n
                print(f"  ✅ Fixtures ({season}): {n} rows")

            elif stype in ("standard", "goalkeeping", "shooting", "playing_time", "misc"):
                n = import_squad_stats(cur, league_id, season_id, stype, headers, data)
                total_squad += n
                print(f"  ✅ Squad {stype} ({season}): {n} rows")

            elif stype == "player":
                n = import_players(cur, season_id, league_id, headers, data)
                total_players += n
                print(f"  ✅ Players ({season}): {n} rows")

        conn.commit()

    print(f"\n🎉 Import complete!")
    print(f"   Fixtures : {total_fixtures}")
    print(f"   Squad    : {total_squad}")
    print(f"   Players  : {total_players}")
    conn.close()

if __name__ == "__main__":
    main()
